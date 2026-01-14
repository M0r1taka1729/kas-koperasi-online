import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime
from fpdf import FPDF

# --- 1. KONFIGURASI & FUNGSI ---
FILE_EXCEL = "database.xlsx"

def export_to_pdf(dataframe):
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, "LAPORAN BUKU KAS KOPERASI", ln=True, align='C')
    pdf.ln(10)
    
    # Lebar kolom disesuaikan dengan kertas Landscape
    widths = [25, 30, 35, 85, 30, 30, 35] 
    headers = ["Bulan", "Tanggal", "No Bukti", "Keterangan", "Debit", "Kredit", "Saldo"]

    pdf.set_fill_color(200, 220, 255)
    pdf.set_font("Arial", 'B', 10)
    for i, h in enumerate(headers):
        pdf.cell(widths[i], 10, h, border=1, fill=True, align='C')
    pdf.ln()
    
    pdf.set_font("Arial", size=9)
    for index, row in dataframe.iterrows():
        fill = index % 2 == 0
        pdf.set_fill_color(245, 245, 245)
        # Ambil data 7 kolom pertama (0 sampai 6)
        for j in range(7):
            val = str(row.iloc[j]) if pd.notnull(row.iloc[j]) else ""
            # Format angka untuk kolom E, F, G (indeks 4, 5, 6)
            if j >= 4:
                try:
                    val_num = float(val)
                    val = f"{val_num:,.0f}"
                except:
                    val = "0"
                pdf.cell(widths[j], 8, val, border=1, fill=fill, align='R')
            else:
                pdf.cell(widths[j], 8, val[:45], border=1, fill=fill)
        pdf.ln()
    return pdf.output(dest='S').encode('latin-1')

# --- 2. TAMPILAN UTAMA ---
st.set_page_config(page_title="Input Kas Koperasi", layout="wide")
st.title("📖 Form Buku Kas Koperasi")

# Form Input
with st.form("form_kas", clear_on_submit=True):
    col1, col2 = st.columns(2)
    with col1:
        tgl = st.date_input("Tanggal", datetime.now())
        no_bukti = st.text_input("No. Bukti")
        keterangan = st.text_input("Keterangan")
    with col2:
        jenis = st.selectbox("Jenis", ["Debit", "Kredit"])
        nominal = st.number_input("Nominal", min_value=0)
    
    submit = st.form_submit_button("Simpan ke Excel")

# Logika Simpan (Hanya dijalankan saat tombol ditekan)
if submit:
    if os.path.exists(FILE_EXCEL):
        wb = load_workbook(FILE_EXCEL)
        ws = wb.active
        row = 7
        while ws[f"B{row}"].value is not None: row += 1
        
        ws[f"A{row}"] = tgl.strftime("%B")
        ws[f"B{row}"] = tgl.strftime("%d/%m/%Y")
        ws[f"C{row}"] = no_bukti
        ws[f"D{row}"] = keterangan
        ws[f"E{row}"] = nominal if jenis == "Debit" else 0
        ws[f"F{row}"] = nominal if jenis == "Kredit" else 0
        
        saldo_ats = ws[f"G{row-1}"].value if row > 7 else 0
        if not isinstance(saldo_ats, (int, float)): saldo_ats = 0
        ws[f"G{row}"] = saldo_ats + ws[f"E{row}"].value - ws[f"F{row}"].value
        
        wb.save(FILE_EXCEL)
        st.success("Data Berhasil Disimpan!")
        st.rerun() # Refresh agar tabel di bawah langsung update

# --- 3. BAGIAN DOWNLOAD (DI LUAR FORM - AGAR SELALU MUNCUL) ---
st.divider()
st.subheader("📥 Cetak & Unduh Laporan")

if os.path.exists(FILE_EXCEL):
    # Membaca data untuk preview dan PDF
    # Gunakan header=None karena kita mulai dari baris 7 (data murni)
    df_data = pd.read_excel(FILE_EXCEL, skiprows=6, header=None)
    
    if not df_data.empty:
        col_pdf, col_xlsx = st.columns(2)
        
        with col_pdf:
            try:
                pdf_output = export_to_pdf(df_data)
                st.download_button(
                    label="📄 Unduh PDF",
                    data=pdf_output,
                    file_name="Laporan_Kas_Koperasi.pdf",
                    mime="application/pdf"
                )
            except Exception as e:
                st.error(f"Gagal menyiapkan PDF: {e}")
        
        with col_xlsx:
            with open(FILE_EXCEL, "rb") as f:
                st.download_button(
                    label="Excel (.xlsx)",
                    data=f,
                    file_name="Buku_Kas_Koperasi.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        st.write("Preview Data Terakhir:")
        st.dataframe(df_data.tail(10))
    else:
        st.info("Belum ada data di baris ke-7 dan seterusnya.")
else:
    st.warning("File database.xlsx belum ditemukan di folder.")
